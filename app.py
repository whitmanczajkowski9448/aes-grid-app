import re
from collections import defaultdict
from copy import copy
from datetime import datetime, date, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo
from urllib.parse import urlparse

import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# APP SETTINGS
# UPDATED: Multi-event support for combined grids and comparisons
# =========================

BASE_URL = "https://results.advancedeventsystems.com/api/event"
MATCH_ENDPOINT_CONSTANT = "240"

TEMPLATE_PATH = "GRID EXAMPLE.xlsx"
TIMEZONE = "America/New_York"

ASSIGNMENT_ROLE_OPTIONS = [
    {
        "id": "match",
        "checkbox_label": "Match name",
        "grid_label": "match",
        "template_offset": 0,
    },
    {
        "id": "format",
        "checkbox_label": "Format",
        "grid_label": "format",
        "template_offset": 1,
    },
    {
        "id": "r1",
        "checkbox_label": "R1",
        "grid_label": "R1",
        "template_offset": 2,
    },
    {
        "id": "r2",
        "checkbox_label": "R2",
        "grid_label": "R2",
        "template_offset": 3,
    },
    {
        "id": "lj1",
        "checkbox_label": "LJ 1",
        "grid_label": "LJ",
        "template_offset": 4,
    },
    {
        "id": "lj2",
        "checkbox_label": "LJ 2",
        "grid_label": "LJ",
        "template_offset": 5,
    },
    {
        "id": "sk",
        "checkbox_label": "SK",
        "grid_label": "SK",
        "template_offset": 6,
    },
    {
        "id": "as",
        "checkbox_label": "AS",
        "grid_label": "AS",
        "template_offset": 7,
    },
]

DEFAULT_ASSIGNMENT_ROLE_IDS = [option["id"] for option in ASSIGNMENT_ROLE_OPTIONS]
ROLE_ROWS = [option["grid_label"] for option in ASSIGNMENT_ROLE_OPTIONS]


def assignment_role_options_from_ids(selected_role_ids: list[str] | None = None) -> list[dict]:
    """Return assignment-grid row options in the normal row order."""
    if selected_role_ids is None:
        selected_role_ids = DEFAULT_ASSIGNMENT_ROLE_IDS

    selected_ids = set(selected_role_ids)
    selected_options = [
        option
        for option in ASSIGNMENT_ROLE_OPTIONS
        if option["id"] in selected_ids
    ]

    # Always return at least one row so the workbook structure remains valid.
    return selected_options or [ASSIGNMENT_ROLE_OPTIONS[0]]

WEEKDAY_ABBREVIATIONS = {
    0: "MO",
    1: "TU",
    2: "WE",
    3: "TH",
    4: "FR",
    5: "SA",
    6: "SU",
}

DEFAULT_LEVEL_ABBREVIATIONS = {
    "OPEN": "O",
    "PREMIER": "P",
    "ELITE": "E",
    "SELECT": "S",
    "ASCEND": "A",
    "CLUB": "C",
    "ASPIRE": "AS",
    "SPIRIT": "SP",
    "CLASSIC": "CL",
    "GIRLS": "U",
    "GIRL": "U",
    "POWER": "P",
    "P": "P",
    "GOLD": "G",
    "SILVER": "S",
    "BRONZE": "B",
    "PLATINUM": "P",
    "DIAMOND": "D",
    "RUBY": "R",
    "SAPPHIRE": "S",
    "EMERALD": "E",
    "NATIONAL": "N",
    "AMERICAN": "A",
    "REGIONAL": "R",
    "USA": "U",
}


# =========================
# QUICK EVENT BUTTONS
# =========================
# Replace the url values below with full AES schedule URLs or raw AES event keys.
# Leave url blank or set active to False to gray out that button.
PRESET_EVENT_BUTTON_ROWS = [
    [
        {"label": "AAU Wave 1", "url": "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjY90/home?_gl=1*11ee7ca*_ga*MjA1MjMzMTg2Ny4xNzY2NzgwNjUz*_ga_PQ25JN9PJ8*czE3ODE2NTM3NjkkbzkyJGcxJHQxNzgxNjUzOTYxJGo0MCRsMCRoMA..&_ga=2.65442030.1738899225.1781380090-2052331867.1766780653", "active": False},
        {"label": "AAU Wave 2", "url": "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjc90/home?_gl=1*11ei72s*_ga*MjA1MjMzMTg2Ny4xNzY2NzgwNjUz*_ga_PQ25JN9PJ8*czE3ODE2NTM3NjkkbzkyJGcxJHQxNzgxNjUzOTc5JGoyMiRsMCRoMA..&_ga=2.68547665.1738899225.1781380090-2052331867.1766780653", "active": False},
        {"label": "AAU Wave 3", "url": "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjg90/court-schedule", "active": False},
        {"label": "AAU Wave 4", "url": "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjk90/court-schedule", "active": True},
        {"label": "AAU Wave 5", "url": "", "active": False},
        {"label": "AAU Wave 6", "url": "", "active": False},
    ],
    [
        {
            "label": "AAU Wave 5 (Boys and Girls)",
            "urls": [
                "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjg90/court-schedule",
                "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjk90/court-schedule",
            ],
            "active": False,
        },
        {
            "label": "AAU Wave 6 (Boys and Girls)",
            "urls": [
                "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjg90/court-schedule",
                "https://results.advancedeventsystems.com/event/PTAwMDAwNDUwMjk90/court-schedule",
            ],
            "active": False,
        },
        {"label": "USAV GJNC 14-17", "url": "https://results.advancedeventsystems.com/event/PTAwMDAwNDIwNjI90/court-schedule", "active": True},
        {"label": "USAV BJNC", "url": "", "active": False},
    ],
]



# =========================
# INPUT HELPERS
# =========================

def extract_event_key(user_input: str) -> str:
    text = (user_input or "").strip()

    if not text:
        return ""

    if text.startswith("http://") or text.startswith("https://"):
        parsed = urlparse(text)
        path_parts = [part for part in parsed.path.split("/") if part]

        if "event" in path_parts:
            event_index = path_parts.index("event")
            if event_index + 1 < len(path_parts):
                return path_parts[event_index + 1].strip()

    match = re.search(r"/event/([^/?#\s]+)", text)
    if match:
        return match.group(1).strip()

    match = re.search(r"/api/event/([^/?#\s]+)", text)
    if match:
        return match.group(1).strip()

    return text


def extract_event_keys(user_input: str) -> list[str]:
    """Return unique AES event keys from pasted keys, URLs, commas, semicolons, or new lines."""
    raw_text = (user_input or "").strip()

    if not raw_text:
        return []

    pieces = [piece.strip() for piece in re.split(r"[\n,;]+", raw_text) if piece.strip()]
    keys = []

    for piece in pieces:
        key = extract_event_key(piece)
        if key:
            keys.append(key)

    return list(dict.fromkeys(keys))


def event_key_input_text(values) -> str:
    """Convert a preset url/url-list/key/key-list value into textarea-friendly input."""
    if values is None:
        return ""

    if isinstance(values, str):
        return values.strip()

    return "\n".join(str(value).strip() for value in values if str(value).strip())


def composite_division_id(event_key: str, division_id) -> str:
    """Avoid division-id collisions when multiple AES events are combined."""
    return f"{event_key}:{division_id}"


def event_records_have_multiple_events(event_records: list[dict]) -> bool:
    return len(event_records or []) > 1


def event_record_key_signature(event_records: list[dict]) -> tuple[str, ...]:
    return tuple(record.get("event_key", "") for record in event_records or [])


def event_record_date_range(record: dict) -> tuple[date, date]:
    info = record.get("event_info", {}) or {}
    return parse_event_date(info["StartDate"]), parse_event_date(info["EndDate"])


def event_record_includes_date(record: dict, event_date: date) -> bool:
    start_date, end_date = event_record_date_range(record)
    return start_date <= event_date <= end_date


def annotate_division_for_event(division: dict, event_key: str) -> dict:
    annotated = dict(division or {})
    annotated["OriginalDivisionId"] = annotated.get("DivisionId")
    annotated["DivisionId"] = composite_division_id(event_key, annotated.get("DivisionId"))
    annotated["SourceEventKey"] = event_key
    return annotated


def annotate_match_for_event(match: dict, event_key: str) -> dict:
    annotated_match = dict(match or {})
    annotated_match["SourceEventKey"] = event_key

    division = dict((match or {}).get("Division", {}) or {})
    if division:
        division["OriginalDivisionId"] = division.get("DivisionId")
        division["DivisionId"] = composite_division_id(event_key, division.get("DivisionId"))
        division["SourceEventKey"] = event_key
        annotated_match["Division"] = division

    return annotated_match


def annotate_match_data_for_event(match_data: dict, event_key: str, event_info: dict | None = None) -> dict:
    """Make a match-data object safe to merge with other events."""
    event_label = (event_info or {}).get("Name", event_key)
    annotated_courts = []

    for court in (match_data or {}).get("CourtSchedules", []) or []:
        annotated_court = dict(court or {})
        annotated_court["SourceEventKey"] = event_key
        annotated_court["SourceEventName"] = event_label
        annotated_court["CourtMatches"] = [
            annotate_match_for_event(match, event_key)
            for match in (court or {}).get("CourtMatches", []) or []
        ]
        annotated_courts.append(annotated_court)

    annotated = dict(match_data or {})
    annotated["CourtSchedules"] = annotated_courts
    return annotated


def merge_match_data_objects(match_data_objects: list[dict]) -> dict:
    """Merge several AES match-data responses into one CourtSchedules structure.

    Courts with the same displayed name are combined. If two event schedules would
    place a match on the same court at the same start time, the later event gets
    its own event-labeled court column so one match does not overwrite another.
    """
    merged_courts = {}
    court_order = []

    def start_time_keys(court: dict) -> set:
        return {
            match.get("ScheduledStartDateTime")
            for match in court.get("CourtMatches", []) or []
            if match.get("ScheduledStartDateTime") is not None
        }

    for match_data in match_data_objects:
        for court in (match_data or {}).get("CourtSchedules", []) or []:
            court_name = str(court.get("Name", "")).strip()
            base_key = normalize_court_exact_key(court_name)
            court_key = base_key

            if court_key in merged_courts:
                existing_starts = start_time_keys(merged_courts[court_key])
                incoming_starts = start_time_keys(court)

                if existing_starts & incoming_starts:
                    source_label = compact_event_name(court.get("SourceEventName", "Event"))
                    court_key = normalize_court_exact_key(f"{source_label}_{court_name}")
                    court = dict(court)
                    court["Name"] = f"{court.get('SourceEventName', 'Event')} - {court_name}"

            if court_key not in merged_courts:
                merged_court = dict(court)
                merged_court["CourtMatches"] = []
                merged_courts[court_key] = merged_court
                court_order.append(court_key)

            merged_courts[court_key]["CourtMatches"].extend(
                list(court.get("CourtMatches", []) or [])
            )

    for court in merged_courts.values():
        court["CourtMatches"] = sorted(
            court.get("CourtMatches", []) or [],
            key=lambda match: (
                match.get("ScheduledStartDateTime", 0),
                match.get("ScheduledEndDateTime", 0),
                clean_match_name(match.get("CompleteShortName", "")),
            ),
        )

    ordered_courts = [merged_courts[key] for key in court_order]

    return {"CourtSchedules": ordered_courts}


def build_event_records(event_keys: list[str]) -> list[dict]:
    records = []

    for event_key in event_keys:
        info = get_event_info(event_key)
        records.append({"event_key": event_key, "event_info": info})

    return records


def build_combined_event_info(event_records: list[dict]) -> dict:
    """Return a single event_info-like dict for one event or an event set."""
    if not event_records:
        return {}

    if len(event_records) == 1:
        record = event_records[0]
        info = dict(record["event_info"] or {})
        event_key = record["event_key"]
        info["Divisions"] = [
            annotate_division_for_event(division, event_key)
            for division in info.get("Divisions", []) or []
        ]
        info["EventKeys"] = [event_key]
        info["Events"] = [{"event_key": event_key, "event_info": record["event_info"]}]
        return info

    starts = []
    ends = []
    names = []
    locations = []
    divisions = []

    for record in event_records:
        event_key = record["event_key"]
        info = record["event_info"] or {}
        starts.append(parse_event_date(info["StartDate"]))
        ends.append(parse_event_date(info["EndDate"]))
        names.append(info.get("Name") or event_key)

        location = (info.get("Location") or "").strip()
        if location and location not in locations:
            locations.append(location)

        divisions.extend(
            annotate_division_for_event(division, event_key)
            for division in info.get("Divisions", []) or []
        )

    return {
        "Name": f"Combined_{len(event_records)}_Events",
        "DisplayName": "Combined Events",
        "CombinedEventNames": names,
        "Location": "; ".join(locations),
        "StartDate": min(starts).isoformat(),
        "EndDate": max(ends).isoformat(),
        "Divisions": divisions,
        "EventKeys": [record["event_key"] for record in event_records],
        "Events": event_records,
    }


def get_combined_match_info(event_records: list[dict], event_date: date) -> dict:
    """Fetch and merge match data for every event in the set that contains event_date."""
    match_data_objects = []

    for record in event_records or []:
        if not event_record_includes_date(record, event_date):
            continue

        event_key = record["event_key"]
        event_info = record.get("event_info", {}) or {}
        match_data = get_match_info(event_key, event_date)
        match_data_objects.append(
            annotate_match_data_for_event(
                match_data=match_data,
                event_key=event_key,
                event_info=event_info,
            )
        )

    return merge_match_data_objects(match_data_objects)


def clear_assignment_download_state():
    keys_to_clear = [
        "assignment_download_custom",
        "show_assignment_download_options",
    ]

    for key in list(st.session_state.keys()):
        if key.startswith("assignment_row_include_") or key in keys_to_clear:
            del st.session_state[key]


def reset_app():
    clear_assignment_download_state()

    keys_to_clear = [
        "event_info",
        "event_key",
        "event_keys",
        "event_records",
        "event_input_value",
        "generated_workbooks",
        "level_rows",
        "level_editor",
        "event_match_counts",
        "comparison_result",
        "comparison_uploaded_summary",
        "comparison_selected_date",
        "comparison_changed_workbooks",
    ]

    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]



# =========================
# API HELPERS
# =========================

def fetch_json(url: str) -> dict:
    headers = {
        "accept": "application/json, text/plain, */*",
        "user-agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
    }

    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.json()


def get_event_info(event_key: str) -> dict:
    return fetch_json(f"{BASE_URL}/{event_key}")


def get_match_info(event_key: str, event_date: date) -> dict:
    date_string = event_date.strftime("%Y-%m-%d")
    return fetch_json(
        f"{BASE_URL}/{event_key}/courts/{date_string}/{MATCH_ENDPOINT_CONSTANT}"
    )


# =========================
# DATE / TIME HELPERS
# =========================

def parse_event_date(value: str) -> date:
    return datetime.fromisoformat(value.split(".")[0]).date()


def date_range(start_date: date, end_date: date):
    current = start_date

    while current <= end_date:
        yield current
        current += timedelta(days=1)


def format_long_date(d: date) -> str:
    return f"{d.strftime('%A')}, {d.strftime('%B')} {d.day}"


def sheet_name_for_date(d: date) -> str:
    weekday = WEEKDAY_ABBREVIATIONS[d.weekday()]
    return f"{weekday} {d.month}-{d.day}"


def c1_label_for_date(d: date) -> str:
    weekday = WEEKDAY_ABBREVIATIONS[d.weekday()]
    return f"{weekday} {d.month}/{d.day}"


def ms_to_local_datetime(ms: int, tz_name: str) -> datetime:
    tz = ZoneInfo(tz_name)
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone(tz)


def timestamp_suffix(tz_name: str) -> str:
    now = datetime.now(ZoneInfo(tz_name))
    return now.strftime("__%y%m%d_%H%M")


def round_up_to_half_hour(dt: datetime) -> datetime:
    dt = dt.replace(second=0, microsecond=0)

    if dt.minute == 0 or dt.minute == 30:
        return dt

    if dt.minute < 30:
        return dt.replace(minute=30)

    return dt.replace(minute=0) + timedelta(hours=1)


def round_match_start_to_next_grid_time(match_start: datetime, grid_times: list[datetime]) -> datetime | None:
    clean_start = match_start.replace(second=0, microsecond=0)

    for grid_time in grid_times:
        if grid_time >= clean_start:
            return grid_time

    return None


def match_duration_minutes(match: dict) -> int:
    start_ms = match["ScheduledStartDateTime"]
    end_ms = match["ScheduledEndDateTime"]

    return round((end_ms - start_ms + 1) / 1000 / 60)


def determine_match_format(match: dict) -> str:
    minutes = match_duration_minutes(match)

    if minutes < 60:
        return "1 set to 25"

    if minutes == 60:
        return "best 2/3"

    return "auto-3"


# =========================
# DIVISION / MATCH NAME HELPERS
# =========================

def normalize_power_text(text: str) -> str:
    """Normalize common words that should become compact division-code text."""
    if not text:
        return ""

    text = str(text)

    # AES sometimes uses words that should become code letters in our grids.
    text = re.sub(r"\bPower\b", "P", text, flags=re.IGNORECASE)
    text = re.sub(r"Power", "P", text, flags=re.IGNORECASE)

    return text


def normalize_level_name(level_name: str) -> str:
    level_name = normalize_power_text((level_name or "").strip())
    level_name = re.sub(r"\s+", " ", level_name)
    return level_name


def normalize_hex_color(hex_color: str | None) -> str:
    if not hex_color:
        return "FFFFFFFF"

    clean = str(hex_color).strip().replace("#", "").upper()

    if len(clean) == 6:
        return f"FF{clean}"

    if len(clean) == 8:
        return clean

    return "FFFFFFFF"


def display_hex_color(hex_color: str | None) -> str:
    if not hex_color:
        return "#FFFFFF"

    clean = str(hex_color).strip().replace("#", "").upper()

    if len(clean) == 6:
        return f"#{clean}"

    if len(clean) == 8:
        return f"#{clean[-6:]}"

    return "#FFFFFF"


def division_age(division_name: str) -> str:
    match = re.search(r"^\s*(\d{1,2})", division_name or "")
    return match.group(1) if match else ""


def division_level_name(division_name: str) -> str:
    name = normalize_power_text((division_name or "").strip())
    name = re.sub(r"^\d{1,2}\s*", "", name).strip()
    name = normalize_level_name(name)

    return name or "Division"


def default_level_abbreviation(level_name: str) -> str:
    level = normalize_level_name(level_name)

    if not level:
        return ""

    lookup_key = level.upper()

    if lookup_key in DEFAULT_LEVEL_ABBREVIATIONS:
        return DEFAULT_LEVEL_ABBREVIATIONS[lookup_key]

    words = [word for word in re.split(r"\s+", level.strip()) if word]

    if not words:
        return ""

    # If a division level starts with Girls/Girl, make that part U.
    # Examples: Girls -> U, Girls Classic -> UCL, Girls Open -> UO.
    first_word_key = words[0].upper()
    if first_word_key in {"GIRLS", "GIRL"}:
        if len(words) == 1:
            return "U"
        return "U" + default_level_abbreviation(" ".join(words[1:]))

    return words[0][0].upper()


def clean_match_name(raw_name: str) -> str:
    """
    Final match-name cleaner.

    Rules:
    - Convert Girls/Girl to U.
    - Remove R1, R2, R3, R4, R5, etc. anywhere in the match name.
    - Delete all lowercase letters.
    - Delete all symbols/spaces/punctuation.
    - Keep only numbers and capital letters.

    Examples:
    14CR1AM1 -> 14CAM1
    14Cr1aM1 -> 14CM1
    15PowerR2BM3 -> 15PBM3
    18 Girls M1 -> 18UM1
    """
    if not raw_name:
        return ""

    text = normalize_power_text(str(raw_name))
    text = re.sub(r"\bGirls\b", "U", text, flags=re.IGNORECASE)
    text = re.sub(r"\bGirl\b", "U", text, flags=re.IGNORECASE)

    # Remove round labels like R1, R2, R10 anywhere in the name.
    text = re.sub(r"R\d+", "", text)

    # Keep only numeric and capital characters.
    text = re.sub(r"[^A-Z0-9]", "", text)

    return text


def build_level_rows(divisions: list[dict]) -> list[dict]:
    level_map = {}

    for division in divisions:
        division_name = division.get("Name", "")
        level = division_level_name(division_name)

        if level not in level_map:
            level_map[level] = {
                "Division Level": level,
                "Abbreviation": default_level_abbreviation(level),
                "Divisions Using This Level": [],
            }

        level_map[level]["Divisions Using This Level"].append(division_name)

    rows = []

    for level in sorted(level_map.keys()):
        row = level_map[level]
        row["Divisions Using This Level"] = ", ".join(row["Divisions Using This Level"])
        rows.append(row)

    return rows


def resolve_level_abbreviations(edited_level_rows: list[dict]) -> dict:
    level_abbreviations = {}

    for row in edited_level_rows:
        level = normalize_level_name(str(row.get("Division Level", "")).strip())
        abbreviation = normalize_power_text(str(row.get("Abbreviation", "")).strip())

        if not abbreviation:
            abbreviation = default_level_abbreviation(level)

        # Abbreviations should only use capital letters/numbers.
        abbreviation = re.sub(r"[^A-Z0-9]", "", abbreviation)

        level_abbreviations[level] = abbreviation

    return level_abbreviations


def build_division_settings_from_levels(divisions: list[dict], level_abbreviations: dict) -> dict:
    division_settings = {}

    for division in divisions:
        division_id = division.get("DivisionId")
        division_name = division.get("Name", "")
        age = division_age(division_name)
        level = division_level_name(division_name)

        level_abbreviation = level_abbreviations.get(
            level,
            default_level_abbreviation(level),
        )

        final_abbreviation = f"{age}{level_abbreviation}" if age else level_abbreviation

        division_settings[division_id] = {
            "abbreviation": clean_match_name(final_abbreviation),
            "color": display_hex_color(division.get("ColorHex")),
        }

    return division_settings


def get_division_correction(match: dict, division_settings: dict) -> dict:
    division = match.get("Division", {}) or {}
    division_id = division.get("DivisionId")

    if division_id in division_settings:
        return division_settings[division_id]

    division_name = division.get("Name", "")
    age = division_age(division_name)
    level = division_level_name(division_name)
    level_abbreviation = default_level_abbreviation(level)
    final_abbreviation = f"{age}{level_abbreviation}" if age else level_abbreviation

    return {
        "abbreviation": clean_match_name(final_abbreviation),
        "color": display_hex_color(division.get("ColorHex")),
    }


def build_assignment_match_display_name(match: dict, division_settings: dict) -> str:
    correction = get_division_correction(match, division_settings)
    division_code = correction["abbreviation"]

    short_name = match.get("CompleteShortName") or ""

    return clean_match_name(f"{division_code}{short_name}")


def build_worksheet_match_display_name(match: dict, division_settings: dict) -> str:
    correction = get_division_correction(match, division_settings)
    division_code = correction["abbreviation"]

    short_name = match.get("CompleteShortName") or ""

    return clean_match_name(f"{division_code}{short_name}")


# =========================
# WORKBOOK STYLE HELPERS
# =========================

def compact_event_name(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", text or "AESEvent")


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)

    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def bolded_font_from_cell(cell):
    existing = copy(cell.font)
    existing.bold = True
    return existing


def load_template_sheet_and_theme(template_path: str | None):
    if not template_path:
        return None, None

    path = Path(template_path)

    if not path.exists():
        return None, None

    template_wb = load_workbook(path)
    template_ws = template_wb[template_wb.sheetnames[0]]
    template_theme = template_wb.loaded_theme

    return template_ws, template_theme


def template_row_for_output_row(
    output_row: int,
    assignment_role_options: list[dict] | None = None,
) -> int:
    if output_row == 1:
        return 1

    role_options = assignment_role_options or ASSIGNMENT_ROLE_OPTIONS
    role_index = (output_row - 2) % len(role_options)
    template_offset = role_options[role_index].get("template_offset", role_index)

    return 2 + template_offset


def clear_unused_match_cell_fills(ws, start_row: int, max_row: int, start_col: int, max_col: int):
    if max_col < start_col:
        return

    for row in range(start_row, max_row + 1):
        for col in range(start_col, max_col + 1):
            cell = ws.cell(row=row, column=col)

            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = PatternFill(fill_type=None)


def apply_template_columns_a_to_d(
    ws,
    template_ws,
    max_row: int,
    assignment_role_options: list[dict] | None = None,
):
    if template_ws is None:
        apply_builtin_columns_a_to_d_style(ws, max_row)
        return

    for col in range(1, 5):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width

    for row in range(1, max_row + 1):
        template_row = template_row_for_output_row(row, assignment_role_options)

        if template_ws.row_dimensions[template_row].height is not None:
            ws.row_dimensions[row].height = template_ws.row_dimensions[template_row].height

        for col in range(1, 5):
            copy_cell_style(
                template_ws.cell(row=template_row, column=col),
                ws.cell(row=row, column=col),
            )

    for row in range(1, max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        c_cell.number_format = "h:mm AM/PM"

        if c_cell.value is not None:
            c_cell.font = bolded_font_from_cell(c_cell)

    d1 = ws["D1"]
    d1.value = None
    d1.fill = PatternFill(fill_type=None)


def apply_builtin_columns_a_to_d_style(ws, max_row: int):
    fallback_blue = "FFD9EAF7"

    widths = {
        "A": 10.15625,
        "B": 12.41796875,
        "C": 8.578125,
        "D": 7.0,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, max_row + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        c_cell = ws.cell(row=row, column=3)
        c_cell.number_format = "h:mm AM/PM"

        if c_cell.value is not None:
            c_cell.font = Font(name="Calibri", size=11, bold=True)

        d_cell = ws.cell(row=row, column=4)
        d_cell.fill = PatternFill(fill_type="solid", fgColor=fallback_blue)
        d_cell.font = Font(name="Calibri", size=11, bold=True)
        d_cell.alignment = Alignment(horizontal="center", vertical="center")

    d1 = ws["D1"]
    d1.value = None
    d1.fill = PatternFill(fill_type=None)


def apply_template_court_header_style(ws, template_ws, start_col: int, end_col: int):
    if template_ws is None:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = 15.68359375
        return

    source_cell = template_ws["E1"]
    source_width = template_ws.column_dimensions["E"].width

    for col in range(start_col, end_col + 1):
        copy_cell_style(source_cell, ws.cell(row=1, column=col))
        ws.column_dimensions[get_column_letter(col)].width = source_width


def apply_template_match_area_style(
    ws,
    template_ws,
    start_col: int,
    end_col: int,
    max_row: int,
    assignment_role_options: list[dict] | None = None,
):
    if template_ws is None:
        for row in range(2, max_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).font = Font(name="Calibri", size=11)
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
        return

    for row in range(2, max_row + 1):
        template_row = template_row_for_output_row(row, assignment_role_options)

        for col in range(start_col, end_col + 1):
            copy_cell_style(
                template_ws.cell(row=template_row, column=5),
                ws.cell(row=row, column=col),
            )


# =========================
# DATA SHAPING
# =========================

def flatten_matches(match_data: dict, tz_name: str):
    court_schedules = match_data.get("CourtSchedules", [])

    matches = []
    court_names = []

    for court in court_schedules:
        court_name = court.get("Name", "")
        court_names.append(court_name)

        for match in court.get("CourtMatches", []):
            matches.append(
                {
                    "court_name": court_name,
                    "match": match,
                    "start_dt": ms_to_local_datetime(
                        match["ScheduledStartDateTime"], tz_name
                    ),
                    "end_dt": ms_to_local_datetime(
                        match["ScheduledEndDateTime"], tz_name
                    ),
                }
            )

    return court_names, matches


def count_matches_in_match_data(match_data: dict) -> int:
    return sum(
        len(court.get("CourtMatches", []) or [])
        for court in match_data.get("CourtSchedules", []) or []
    )


def build_event_match_count_summary(event_records: list[dict], event_info: dict) -> dict:
    start_date = parse_event_date(event_info["StartDate"])
    end_date = parse_event_date(event_info["EndDate"])

    by_day = []
    total = 0

    for event_date in date_range(start_date, end_date):
        match_data = get_combined_match_info(event_records, event_date)
        count = count_matches_in_match_data(match_data)
        total += count
        by_day.append({"date": event_date, "count": count})

    return {"by_day": by_day, "total": total}


def render_event_match_counts(match_counts: dict):
    if not match_counts:
        return

    st.metric("Total Matches", match_counts.get("total", 0))

    day_rows = [
        {
            "Day": format_long_date(row["date"]),
            "Matches": row["count"],
        }
        for row in match_counts.get("by_day", [])
    ]

    if day_rows:
        st.markdown("**Match Count by Day**")
        st.table(day_rows)


def get_grid_start_end(sheet_date: date, matches: list, tz_name: str):
    if matches:
        earliest = min(item["start_dt"] for item in matches)
        latest = max(item["end_dt"] for item in matches)

        grid_start = earliest.replace(second=0, microsecond=0)
        grid_end = round_up_to_half_hour(latest)
    else:
        grid_start = datetime(
            sheet_date.year,
            sheet_date.month,
            sheet_date.day,
            8,
            0,
            tzinfo=ZoneInfo(tz_name),
        )
        grid_end = grid_start + timedelta(hours=12)

    return grid_start, grid_end


def build_grid_times(grid_start: datetime, grid_end: datetime) -> list[datetime]:
    grid_times = []
    current_time = grid_start

    while current_time <= grid_end:
        grid_times.append(current_time.replace(second=0, microsecond=0))
        current_time += timedelta(minutes=30)

    return grid_times


def find_time_row_for_match(match_start_dt: datetime, time_lookup: dict):
    clean_start = match_start_dt.replace(second=0, microsecond=0)

    if clean_start in time_lookup:
        return time_lookup[clean_start]

    grid_times = sorted(time_lookup.keys())
    target_time = round_match_start_to_next_grid_time(clean_start, grid_times)

    if target_time is None:
        return None

    return time_lookup[target_time]


# =========================
# ASSIGNMENT GRID WORKBOOK
# =========================

def build_assignment_time_grid(
    ws,
    sheet_date: date,
    grid_start: datetime,
    grid_end: datetime,
    assignment_role_options: list[dict] | None = None,
):
    role_options = assignment_role_options or ASSIGNMENT_ROLE_OPTIONS

    ws["A1"] = "Notes"
    ws["B1"] = "Notes"
    ws["C1"] = c1_label_for_date(sheet_date)
    ws["D1"] = None

    role_output_offsets = {
        option["id"]: offset
        for offset, option in enumerate(role_options)
    }

    time_to_base_row = {}
    grid_times = build_grid_times(grid_start, grid_end)

    row = 2
    for grid_time in grid_times:
        time_to_base_row[grid_time] = row

        time_cell = ws.cell(row=row, column=3)
        time_cell.value = grid_time.time()
        time_cell.number_format = "h:mm AM/PM"

        for offset, option in enumerate(role_options):
            ws.cell(row=row + offset, column=4).value = option["grid_label"]

        row += len(role_options)

    return time_to_base_row, row - 1, role_output_offsets


def build_assignment_grid_sheet(
    ws,
    sheet_date: date,
    match_data: dict,
    tz_name: str,
    division_settings: dict,
    template_ws=None,
    assignment_role_options: list[dict] | None = None,
):
    role_options = assignment_role_options or ASSIGNMENT_ROLE_OPTIONS
    court_names, matches = flatten_matches(match_data, tz_name)
    grid_start, grid_end = get_grid_start_end(sheet_date, matches, tz_name)

    time_to_base_row, max_row, role_output_offsets = build_assignment_time_grid(
        ws,
        sheet_date,
        grid_start,
        grid_end,
        role_options,
    )

    start_court_col = 5
    end_court_col = start_court_col + len(court_names) - 1

    apply_template_columns_a_to_d(ws, template_ws, max_row, role_options)

    if court_names:
        apply_template_court_header_style(ws, template_ws, start_court_col, end_court_col)
        apply_template_match_area_style(
            ws,
            template_ws,
            start_court_col,
            end_court_col,
            max_row,
            role_options,
        )

    for idx, court_name in enumerate(court_names):
        col = start_court_col + idx
        ws.cell(row=1, column=col).value = court_name

    court_to_col = {
        court_name: start_court_col + idx for idx, court_name in enumerate(court_names)
    }

    for item in matches:
        court_name = item["court_name"]
        match = item["match"]

        base_row = find_time_row_for_match(item["start_dt"], time_to_base_row)

        if base_row is None:
            continue

        col = court_to_col[court_name]

        match_name = build_assignment_match_display_name(match, division_settings)

        if match_name:
            correction = get_division_correction(match, division_settings)
            fill_color = normalize_hex_color(correction["color"])

            if "match" in role_output_offsets:
                match_cell = ws.cell(
                    row=base_row + role_output_offsets["match"],
                    column=col,
                )
                match_cell.value = match_name
                match_cell.font = Font(name="Calibri", size=11, bold=True)
                match_cell.alignment = Alignment(horizontal="center", vertical="center")
                match_cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

            if "format" in role_output_offsets:
                format_cell = ws.cell(
                    row=base_row + role_output_offsets["format"],
                    column=col,
                )
                format_cell.value = determine_match_format(match)
                format_cell.font = Font(name="Calibri", size=11, bold=False)
                format_cell.alignment = Alignment(horizontal="center", vertical="center")

    if court_names:
        clear_unused_match_cell_fills(
            ws=ws,
            start_row=2,
            max_row=max_row,
            start_col=start_court_col,
            max_col=end_court_col,
        )

    ws["D1"].value = None
    ws["D1"].fill = PatternFill(fill_type=None)
    ws.freeze_panes = None


# =========================
# WORKSHEET GRID WORKBOOK
# =========================

def build_worksheet_grid_sheet(
    ws,
    sheet_date: date,
    match_data: dict,
    tz_name: str,
    division_settings: dict,
):
    court_names, matches = flatten_matches(match_data, tz_name)
    grid_start, grid_end = get_grid_start_end(sheet_date, matches, tz_name)

    ws["A1"] = c1_label_for_date(sheet_date)
    ws["A1"].font = Font(name="Calibri", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    start_court_col = 2

    for idx, court_name in enumerate(court_names):
        col = start_court_col + idx
        cell = ws.cell(row=1, column=col)
        cell.value = court_name
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    time_to_row = {}
    grid_times = build_grid_times(grid_start, grid_end)

    row = 2
    for grid_time in grid_times:
        time_to_row[grid_time] = row

        time_cell = ws.cell(row=row, column=1)
        time_cell.value = grid_time.time()
        time_cell.number_format = "h:mm AM/PM"
        time_cell.font = Font(name="Calibri", size=11, bold=True)
        time_cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    max_row = row - 1
    max_col = start_court_col + len(court_names) - 1

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=11, bold=(r == 1 or c == 1))

    court_to_col = {
        court_name: start_court_col + idx for idx, court_name in enumerate(court_names)
    }

    for item in matches:
        court_name = item["court_name"]
        match = item["match"]

        row = find_time_row_for_match(item["start_dt"], time_to_row)

        if row is None:
            continue

        col = court_to_col[court_name]
        cell = ws.cell(row=row, column=col)

        match_name = build_worksheet_match_display_name(match, division_settings)

        if match_name:
            cell.value = match_name
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            correction = get_division_correction(match, division_settings)
            fill_color = normalize_hex_color(correction["color"])
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

    if court_names:
        clear_unused_match_cell_fills(
            ws=ws,
            start_row=2,
            max_row=max_row,
            start_col=start_court_col,
            max_col=max_col,
        )

    ws.column_dimensions["A"].width = 10

    for col in range(start_court_col, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "B2"




# =========================
# GRID COMPARISON HELPERS
# =========================

def is_blank_cell_value(value) -> bool:
    return value is None or str(value).strip() == ""


def parse_grid_time_value(value) -> time | None:
    """Parse a worksheet time cell only when it is truly a time value."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)

    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)

    if isinstance(value, (int, float)):
        # Excel stores times as fractional days. Ignore larger values because
        # they are usually dates/counts in this type of workbook.
        if 0 <= value < 1:
            total_minutes = round(value * 24 * 60)
            hour = (total_minutes // 60) % 24
            minute = total_minutes % 60
            return time(hour=hour, minute=minute)
        return None

    if isinstance(value, str):
        text = value.strip()
        match = re.fullmatch(r"(\d{1,2}):(\d{2})(?:\s*([AP]M))?", text, flags=re.IGNORECASE)

        if not match:
            return None

        hour = int(match.group(1))
        minute = int(match.group(2))
        suffix = match.group(3).upper() if match.group(3) else None

        if minute > 59 or hour > 23:
            return None

        if suffix:
            if hour < 1 or hour > 12:
                return None
            if suffix == "PM" and hour != 12:
                hour += 12
            if suffix == "AM" and hour == 12:
                hour = 0

        return time(hour=hour, minute=minute)

    return None


def format_time_key(t: time) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"


def format_time_label(t: time) -> str:
    hour = t.hour % 12 or 12
    suffix = "AM" if t.hour < 12 else "PM"
    return f"{hour}:{t.minute:02d} {suffix}"


def schedule_header_label(value) -> bool:
    if not isinstance(value, str):
        return False

    text = value.strip().upper()
    return bool(
        re.fullmatch(r"[A-Z]{2}\s+\d{1,2}/\d{1,2}", text)
        or re.fullmatch(r"[A-Z]{2}\s+\d{1,2}-\d{1,2}", text)
    )


def find_schedule_header_row(ws) -> int | None:
    """
    Detects the row containing the date/court headers.

    Compatible with:
    - Generated worksheet grids where the schedule starts on row 1.
    - Workbooks with assignment/name rows above the schedule grid.
    """
    max_scan_row = min(ws.max_row, 75)

    for row in range(1, max_scan_row + 1):
        header_values = [ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1)]
        non_empty_headers = [value for value in header_values if not is_blank_cell_value(value)]
        text_header_count = sum(1 for value in non_empty_headers if isinstance(value, str) and value.strip())

        if len(non_empty_headers) < 3:
            continue

        # Most compatible grids have court names as text. If the date label in
        # column A is clear, allow numeric court names too.
        if text_header_count < 3 and not schedule_header_label(ws.cell(row=row, column=1).value):
            continue

        time_offsets_below = []
        for check_row in range(row + 1, min(ws.max_row, row + 6) + 1):
            if parse_grid_time_value(ws.cell(row=check_row, column=1).value) is not None:
                time_offsets_below.append(check_row - row)

        if len(time_offsets_below) >= 2 and min(time_offsets_below) <= 2:
            return row

    return None


def normalize_uploaded_match_value(value) -> str:
    if value is None or isinstance(value, (datetime, date, time, int, float)):
        return ""

    text = str(value).strip()

    if not text or parse_grid_time_value(text) is not None:
        return ""

    match_code = clean_match_name(text)

    if len(match_code) < 3:
        return ""

    if not re.search(r"\d", match_code):
        return ""

    return match_code


def normalize_court_exact_key(court_name: str) -> str:
    clean = re.sub(r"[^A-Z0-9]", "", str(court_name or "").upper())
    return clean or "COURT"


def court_number(court_name: str) -> int | None:
    match = re.search(r"(\d+)\s*$", str(court_name or "").strip())
    return int(match.group(1)) if match else None


def match_age_key(match_code: str) -> str:
    """Use only the leading age for comparison, per requested rule."""
    code = clean_match_name(match_code)

    # Volleyball divisions should normally start with a two-digit age.
    match = re.match(r"^(\d{2})", code)
    if match:
        return match.group(1)

    # Fallback for unusual values: first two cleaned characters.
    return code[:2]


def make_schedule_entry(
    court_name: str,
    match_time: time,
    match_code: str,
    source: str,
    division_label: str = "",
):
    clean_code = clean_match_name(match_code)
    age = match_age_key(clean_code)

    return {
        "court_name": str(court_name or "").strip(),
        "court_key": normalize_court_exact_key(court_name),
        "court_number": court_number(court_name),
        "time_key": format_time_key(match_time),
        "time_label": format_time_label(match_time),
        "match_code": clean_code,
        "match_age": age,
        "source": source,
        "division_label": division_label or describe_match_code(clean_code),
    }


def get_uploaded_workbook_sheet_names(uploaded_file) -> list[str]:
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    return wb.sheetnames


def default_uploaded_sheet_index(sheet_names: list[str], selected_date: date) -> int:
    if not sheet_names:
        return 0

    expected_names = {
        sheet_name_for_date(selected_date).upper(),
        c1_label_for_date(selected_date).upper(),
    }

    for idx, sheet_name in enumerate(sheet_names):
        if sheet_name.upper() in expected_names:
            return idx

    return 0


def parse_uploaded_workbook_schedule(uploaded_file, sheet_name: str | None = None) -> dict:
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row = find_schedule_header_row(ws)

    if header_row is None:
        raise ValueError(
            "Could not find a schedule grid. Make sure the uploaded workbook has court names across the top and times down column A."
        )

    court_columns = []

    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value

        if is_blank_cell_value(value):
            continue

        court_columns.append((col, str(value).strip()))

    entries = []

    for row in range(header_row + 1, ws.max_row + 1):
        parsed_time = parse_grid_time_value(ws.cell(row=row, column=1).value)

        if parsed_time is None:
            continue

        for col, court_name in court_columns:
            match_code = normalize_uploaded_match_value(ws.cell(row=row, column=col).value)

            if match_code:
                entries.append(
                    make_schedule_entry(
                        court_name=court_name,
                        match_time=parsed_time,
                        match_code=match_code,
                        source="uploaded",
                    )
                )

    return {
        "sheet_name": ws.title,
        "header_row": header_row,
        "court_count": len(court_columns),
        "match_count": len(entries),
        "entries": entries,
    }


def grid_time_for_online_match(match_start_dt: datetime, grid_times: list[datetime]) -> time:
    rounded_dt = round_match_start_to_next_grid_time(match_start_dt, grid_times)

    if rounded_dt is None:
        rounded_dt = round_up_to_half_hour(match_start_dt)

    return rounded_dt.time().replace(second=0, microsecond=0)


def build_online_schedule_entries(match_data: dict, tz_name: str, division_settings: dict) -> list[dict]:
    court_names, matches = flatten_matches(match_data, tz_name)

    if matches:
        sheet_date = matches[0]["start_dt"].date()
    else:
        sheet_date = datetime.now(ZoneInfo(tz_name)).date()

    grid_start, grid_end = get_grid_start_end(sheet_date, matches, tz_name)
    grid_times = build_grid_times(grid_start, grid_end)

    entries = []

    for item in matches:
        match = item["match"]
        match_code = build_worksheet_match_display_name(match, division_settings)

        if not match_code:
            continue

        division = match.get("Division", {}) or {}
        division_label = division.get("Name", "") or describe_match_code(match_code)
        match_time = grid_time_for_online_match(item["start_dt"], grid_times)

        entries.append(
            make_schedule_entry(
                court_name=item["court_name"],
                match_time=match_time,
                match_code=match_code,
                source="online",
                division_label=division_label,
            )
        )

    return entries


def describe_match_code(match_code: str) -> str:
    """Best-effort human label for workbook-only match codes."""
    code = clean_match_name(match_code)
    age_match = re.match(r"^(\d{1,2})", code)

    if not age_match:
        return code

    age = age_match.group(1)
    remainder = code[len(age):]

    abbreviation_names = {
        "AS": "Aspire",
        "SP": "Spirit",
        "CL": "Classic",
        "O": "Open",
        "P": "Premier",
        "E": "Elite",
        "S": "Select",
        "A": "Ascend",
        "C": "Club",
        "G": "Gold",
        "B": "Bronze",
        "D": "Diamond",
        "R": "Regional",
        "N": "National",
        "U": "USA",
    }

    for abbreviation in sorted(abbreviation_names.keys(), key=len, reverse=True):
        if remainder.startswith(abbreviation):
            return f"{age} {abbreviation_names[abbreviation]}"

    return code


def build_court_aliases(uploaded_entries: list[dict], online_entries: list[dict]) -> tuple[dict, dict, list[str]]:
    """
    Exact court names are preferred. If exact names differ, unique trailing court
    numbers are used as a fallback so OCCC 1 can compare to North 1 when needed.
    """
    uploaded_courts = {}
    online_courts = {}

    for entry in uploaded_entries:
        uploaded_courts.setdefault(entry["court_key"], entry["court_name"])

    for entry in online_entries:
        online_courts.setdefault(entry["court_key"], entry["court_name"])

    aliases = {key: key for key in online_courts.keys()}
    canonical_names = {key: name for key, name in online_courts.items()}
    notes = []

    def unique_number_map(court_lookup: dict) -> dict:
        grouped = defaultdict(list)

        for key, name in court_lookup.items():
            number = court_number(name)
            if number is not None:
                grouped[number].append((key, name))

        return {
            number: values[0]
            for number, values in grouped.items()
            if len(values) == 1
        }

    uploaded_by_number = unique_number_map(uploaded_courts)
    online_by_number = unique_number_map(online_courts)

    for uploaded_key, uploaded_name in uploaded_courts.items():
        if uploaded_key in online_courts:
            aliases[uploaded_key] = uploaded_key
            canonical_names.setdefault(uploaded_key, online_courts[uploaded_key])
            continue

        number = court_number(uploaded_name)

        if number is not None and number in uploaded_by_number and number in online_by_number:
            online_key, online_name = online_by_number[number]
            aliases[uploaded_key] = online_key
            canonical_names[online_key] = online_name
            notes.append(f"Matched uploaded court '{uploaded_name}' to online court '{online_name}' by court number.")
        else:
            aliases[uploaded_key] = uploaded_key
            canonical_names.setdefault(uploaded_key, uploaded_name)

    return aliases, canonical_names, notes


def aligned_position(entry: dict, court_aliases: dict) -> tuple[str, str]:
    return (court_aliases.get(entry["court_key"], entry["court_key"]), entry["time_key"])


def entry_location_text(entry: dict, canonical_court_names: dict | None = None) -> str:
    if canonical_court_names and "aligned_court_key" in entry:
        court_name = canonical_court_names.get(entry["aligned_court_key"], entry["court_name"])
    else:
        court_name = entry["court_name"]

    return f"{court_name} at {entry['time_label']}"


def entries_by_position(entries: list[dict], court_aliases: dict) -> dict:
    lookup = {}

    for entry in entries:
        aligned_key, time_key = aligned_position(entry, court_aliases)
        entry["aligned_court_key"] = aligned_key
        lookup[(aligned_key, time_key)] = entry

    return lookup


def entries_by_match_code(entries: list[dict]) -> dict:
    lookup = defaultdict(list)

    for entry in entries:
        lookup[entry["match_code"]].append(entry)

    return lookup


def group_change_entries_by_court(entries: list[dict], canonical_court_names: dict) -> dict:
    grouped = defaultdict(list)

    for entry in entries:
        court_name = canonical_court_names.get(entry.get("aligned_court_key"), entry["court_name"])
        grouped[court_name].append(entry)

    return dict(sorted(grouped.items(), key=lambda item: natural_sort_key(item[0])))


def natural_sort_key(text: str):
    return [int(part) if part.isdigit() else part.upper() for part in re.split(r"(\d+)", str(text))]


def sorted_entries_by_time(entries: list[dict]) -> list[dict]:
    return sorted(
        entries,
        key=lambda entry: (
            entry.get("time_key", ""),
            natural_sort_key(entry.get("match_age", "")),
            natural_sort_key(entry.get("match_code", "")),
        ),
    )


def compare_schedule_structures(uploaded_entries: list[dict], online_entries: list[dict]) -> dict:
    """
    Compare the uploaded grid to the online schedule by court/time structure.

    Match-name comparison intentionally uses only the age key, which is the first
    two cleaned characters of the match name/code. This means 18CLM1 vs 18PM1
    is treated as unchanged at the same court/time because both are age 18.
    """
    court_aliases, canonical_court_names, alias_notes = build_court_aliases(uploaded_entries, online_entries)

    uploaded_by_position = entries_by_position(uploaded_entries, court_aliases)
    online_by_position = entries_by_position(online_entries, court_aliases)

    uploaded_positions = set(uploaded_by_position.keys())
    online_positions = set(online_by_position.keys())

    changed_positions = []

    for position in sorted(uploaded_positions & online_positions):
        uploaded_entry = uploaded_by_position[position]
        online_entry = online_by_position[position]

        if uploaded_entry.get("match_age") != online_entry.get("match_age"):
            changed_positions.append({
                "position": position,
                "old": uploaded_entry,
                "new": online_entry,
            })

    candidate_removed = [uploaded_by_position[position] for position in sorted(uploaded_positions - online_positions)]
    candidate_added = [online_by_position[position] for position in sorted(online_positions - uploaded_positions)]

    removed_by_age = defaultdict(list)
    added_by_age = defaultdict(list)

    for entry in candidate_removed:
        removed_by_age[entry.get("match_age", "")].append(entry)

    for entry in candidate_added:
        added_by_age[entry.get("match_age", "")].append(entry)

    moved_matches = []
    moved_removed_ids = set()
    moved_added_ids = set()

    for age in sorted(set(removed_by_age.keys()) & set(added_by_age.keys()), key=natural_sort_key):
        old_entries = sorted_entries_by_time(removed_by_age[age])
        new_entries = sorted_entries_by_time(added_by_age[age])
        pair_count = min(len(old_entries), len(new_entries))

        for idx in range(pair_count):
            old_entry = old_entries[idx]
            new_entry = new_entries[idx]
            moved_removed_ids.add(id(old_entry))
            moved_added_ids.add(id(new_entry))
            moved_matches.append({
                "match_age": age,
                "match_code": age,
                "old_entries": [old_entry],
                "new_entries": [new_entry],
                "division_label": f"Age {age}" if age else "Unknown age",
            })

    added_entries = [entry for entry in candidate_added if id(entry) not in moved_added_ids]
    removed_entries = [entry for entry in candidate_removed if id(entry) not in moved_removed_ids]

    added_by_court = group_change_entries_by_court(added_entries, canonical_court_names)
    removed_by_court = group_change_entries_by_court(removed_entries, canonical_court_names)

    is_unchanged = not any([
        added_entries,
        removed_entries,
        moved_matches,
        changed_positions,
    ])

    return {
        "is_unchanged": is_unchanged,
        "added_entries": added_entries,
        "removed_entries": removed_entries,
        "added_by_court": added_by_court,
        "removed_by_court": removed_by_court,
        "moved_matches": moved_matches,
        "changed_positions": changed_positions,
        "canonical_court_names": canonical_court_names,
        "alias_notes": alias_notes,
        "uploaded_match_count": len(uploaded_entries),
        "online_match_count": len(online_entries),
    }


def changed_court_names_from_comparison(result: dict) -> set[str]:
    changed_names = set()
    canonical_court_names = result.get("canonical_court_names", {})

    def add_entry(entry: dict):
        aligned_key = entry.get("aligned_court_key") or entry.get("court_key")
        name = canonical_court_names.get(aligned_key, entry.get("court_name", ""))
        if name:
            changed_names.add(str(name).strip())

    for entry in result.get("added_entries", []):
        add_entry(entry)

    for entry in result.get("removed_entries", []):
        add_entry(entry)

    for move in result.get("moved_matches", []):
        for entry in move.get("old_entries", []):
            add_entry(entry)
        for entry in move.get("new_entries", []):
            add_entry(entry)

    for change in result.get("changed_positions", []):
        add_entry(change.get("old", {}))
        add_entry(change.get("new", {}))

    return changed_names


def format_time_list(entries: list[dict]) -> str:
    times = []
    seen = set()

    for entry in sorted_entries_by_time(entries):
        if entry["time_key"] not in seen:
            seen.add(entry["time_key"])
            times.append(entry["time_label"])

    return ", ".join(times)



def count_word(count: int) -> str:
    """Return a readable count word for the comparison summary."""
    words = {
        0: "zero",
        1: "one",
        2: "two",
        3: "three",
        4: "four",
        5: "five",
        6: "six",
        7: "seven",
        8: "eight",
        9: "nine",
        10: "ten",
        11: "eleven",
        12: "twelve",
    }
    return words.get(count, str(count))


def render_colored_change_line(court_name: str, entries: list[dict], action: str, color: str):
    """Render court structure changes in the requested sentence style."""
    count = len(entries)
    plural = "match" if count == 1 else "matches"
    times = format_time_list(entries)
    st.markdown(
        f'<div style="color:{color}; font-weight:700; margin: 0.15rem 0;">'
        f'{court_name} {action} {count_word(count)} {plural} ({times}).'
        f'</div>',
        unsafe_allow_html=True,
    )


def render_grouped_change_lines(grouped_entries: dict, action: str, color: str):
    for court_name, entries in grouped_entries.items():
        render_colored_change_line(court_name, entries, action, color)


def render_comparison_results(result: dict, uploaded_summary: dict, selected_date: date):
    st.caption(
        f"Compared {uploaded_summary['match_count']} uploaded matches from '{uploaded_summary['sheet_name']}' "
        f"to {result['online_match_count']} online matches for {format_long_date(selected_date)}. "
        "Match-code comparisons use age only."
    )

    if result["alias_notes"]:
        with st.expander("Court-name matching notes"):
            st.write(f"Matched {len(result['alias_notes'])} uploaded courts to online courts by court number.")
            for note in result["alias_notes"][:20]:
                st.write(f"- {note}")
            if len(result["alias_notes"]) > 20:
                st.write(f"- ...and {len(result['alias_notes']) - 20} more.")

    if result["is_unchanged"]:
        st.success("✅ Grid Structure Verified & Unchanged")
        return

    st.warning("Grid structure changes found.")

    # A moved match is a structure change in two places:
    # - the old court/time lost that match slot
    # - the new court/time added that match slot
    # To keep the output language consistent, moved matches are folded into
    # the same court-based lost/added summary instead of being rendered as
    # separate "Age X moved from..." lines.
    moved_old_entries = [
        entry
        for move in result.get("moved_matches", [])
        for entry in move.get("old_entries", [])
    ]
    moved_new_entries = [
        entry
        for move in result.get("moved_matches", [])
        for entry in move.get("new_entries", [])
    ]

    structure_removed_by_court = group_change_entries_by_court(
        result.get("removed_entries", []) + moved_old_entries,
        result["canonical_court_names"],
    )
    structure_added_by_court = group_change_entries_by_court(
        result.get("added_entries", []) + moved_new_entries,
        result["canonical_court_names"],
    )

    if structure_removed_by_court:
        st.markdown("**Lost matches**")
        render_grouped_change_lines(structure_removed_by_court, "lost", "#C00000")

    if structure_added_by_court:
        st.markdown("**Added matches**")
        render_grouped_change_lines(structure_added_by_court, "added", "#0070C0")

    if result.get("moved_matches"):
        with st.expander("Move details"):
            st.write(
                "Moves are included above as lost slots on the original court/time "
                "and added slots on the new court/time."
            )
            for move in result["moved_matches"]:
                old_locations = ", ".join(
                    entry_location_text(entry, result["canonical_court_names"])
                    for entry in sorted_entries_by_time(move["old_entries"])
                )
                new_locations = ", ".join(
                    entry_location_text(entry, result["canonical_court_names"])
                    for entry in sorted_entries_by_time(move["new_entries"])
                )
                st.write(
                    f"- Age {move['match_age']} moved from {old_locations} to {new_locations}."
                )

    if result["changed_positions"]:
        st.markdown("**Changed age at same court and time**")
        for change in result["changed_positions"]:
            old_entry = change["old"]
            new_entry = change["new"]
            court_name = result["canonical_court_names"].get(
                new_entry.get("aligned_court_key"),
                new_entry["court_name"],
            )
            st.write(
                f"- **{court_name} at {new_entry['time_label']}** is now "
                f"**Age {new_entry.get('match_age', '')}**; "
                f"was **Age {old_entry.get('match_age', '')}**."
            )


# =========================
# CREATE WORKBOOKS IN MEMORY# =========================
# CREATE WORKBOOKS IN MEMORY
# =========================

def save_workbook_to_bytes(wb: Workbook) -> bytes:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_assignment_grid_file(
    event_records: list[dict],
    event_info: dict,
    division_settings: dict,
    assignment_role_ids: list[str] | None = None,
    filename_prefix: str = "AssignmentGrid",
):
    event_name = event_info.get("Name", "AESEvent")
    start_date = parse_event_date(event_info["StartDate"])
    end_date = parse_event_date(event_info["EndDate"])
    role_options = assignment_role_options_from_ids(assignment_role_ids)

    template_ws, template_theme = load_template_sheet_and_theme(TEMPLATE_PATH)

    assignment_wb = Workbook()
    assignment_wb.remove(assignment_wb.active)

    if template_theme:
        assignment_wb.loaded_theme = template_theme

    for event_date in date_range(start_date, end_date):
        sheet_name = sheet_name_for_date(event_date)
        match_data = get_combined_match_info(event_records, event_date)

        assignment_ws = assignment_wb.create_sheet(title=sheet_name)
        build_assignment_grid_sheet(
            ws=assignment_ws,
            sheet_date=event_date,
            match_data=match_data,
            tz_name=TIMEZONE,
            division_settings=division_settings,
            template_ws=template_ws,
            assignment_role_options=role_options,
        )

    event_name_no_spaces = compact_event_name(event_name)
    suffix = timestamp_suffix(TIMEZONE)

    return {
        "assignment_filename": f"{filename_prefix}_{event_name_no_spaces}{suffix}.xlsx",
        "assignment_bytes": save_workbook_to_bytes(assignment_wb),
    }


def create_workbooks(
    event_records: list[dict],
    event_info: dict,
    division_settings: dict,
    assignment_role_ids: list[str] | None = None,
):
    event_name = event_info.get("Name", "AESEvent")

    start_date = parse_event_date(event_info["StartDate"])
    end_date = parse_event_date(event_info["EndDate"])
    role_options = assignment_role_options_from_ids(assignment_role_ids)

    template_ws, template_theme = load_template_sheet_and_theme(TEMPLATE_PATH)

    assignment_wb = Workbook()
    assignment_wb.remove(assignment_wb.active)

    worksheet_wb = Workbook()
    worksheet_wb.remove(worksheet_wb.active)

    if template_theme:
        assignment_wb.loaded_theme = template_theme
        worksheet_wb.loaded_theme = template_theme

    for event_date in date_range(start_date, end_date):
        sheet_name = sheet_name_for_date(event_date)
        match_data = get_combined_match_info(event_records, event_date)

        assignment_ws = assignment_wb.create_sheet(title=sheet_name)
        build_assignment_grid_sheet(
            ws=assignment_ws,
            sheet_date=event_date,
            match_data=match_data,
            tz_name=TIMEZONE,
            division_settings=division_settings,
            template_ws=template_ws,
            assignment_role_options=role_options,
        )

        worksheet_ws = worksheet_wb.create_sheet(title=sheet_name)
        build_worksheet_grid_sheet(
            ws=worksheet_ws,
            sheet_date=event_date,
            match_data=match_data,
            tz_name=TIMEZONE,
            division_settings=division_settings,
        )

    event_name_no_spaces = compact_event_name(event_name)
    suffix = timestamp_suffix(TIMEZONE)

    assignment_filename = f"AssignmentGrid_{event_name_no_spaces}{suffix}.xlsx"
    worksheet_filename = f"WorksheetGrid_{event_name_no_spaces}{suffix}.xlsx"

    return {
        "assignment_filename": assignment_filename,
        "assignment_bytes": save_workbook_to_bytes(assignment_wb),
        "worksheet_filename": worksheet_filename,
        "worksheet_bytes": save_workbook_to_bytes(worksheet_wb),
    }


def court_name_is_in_changed_set(court_name: str, changed_court_names: set[str]) -> bool:
    if not changed_court_names:
        return False

    exact_names = {str(name).strip() for name in changed_court_names}
    exact_keys = {normalize_court_exact_key(name) for name in changed_court_names}
    changed_numbers = {court_number(name) for name in changed_court_names}
    changed_numbers.discard(None)

    return (
        str(court_name or "").strip() in exact_names
        or normalize_court_exact_key(court_name) in exact_keys
        or court_number(court_name) in changed_numbers
    )


def filter_match_data_to_courts(match_data: dict, changed_court_names: set[str]) -> dict:
    filtered = dict(match_data or {})
    filtered["CourtSchedules"] = [
        court
        for court in (match_data or {}).get("CourtSchedules", []) or []
        if court_name_is_in_changed_set(court.get("Name", ""), changed_court_names)
    ]
    return filtered


def create_changed_court_workbooks(
    event_info: dict,
    sheet_date: date,
    match_data: dict,
    division_settings: dict,
    changed_court_names: set[str],
):
    event_name = event_info.get("Name", "AESEvent")
    event_name_no_spaces = compact_event_name(event_name)
    suffix = timestamp_suffix(TIMEZONE)
    date_suffix = sheet_date.strftime("%m%d")

    template_ws, template_theme = load_template_sheet_and_theme(TEMPLATE_PATH)
    filtered_match_data = filter_match_data_to_courts(match_data, changed_court_names)

    assignment_wb = Workbook()
    assignment_wb.remove(assignment_wb.active)

    worksheet_wb = Workbook()
    worksheet_wb.remove(worksheet_wb.active)

    if template_theme:
        assignment_wb.loaded_theme = template_theme
        worksheet_wb.loaded_theme = template_theme

    sheet_name = sheet_name_for_date(sheet_date)

    assignment_ws = assignment_wb.create_sheet(title=sheet_name)
    build_assignment_grid_sheet(
        ws=assignment_ws,
        sheet_date=sheet_date,
        match_data=filtered_match_data,
        tz_name=TIMEZONE,
        division_settings=division_settings,
        template_ws=template_ws,
    )

    worksheet_ws = worksheet_wb.create_sheet(title=sheet_name)
    build_worksheet_grid_sheet(
        ws=worksheet_ws,
        sheet_date=sheet_date,
        match_data=filtered_match_data,
        tz_name=TIMEZONE,
        division_settings=division_settings,
    )

    court_count = len(filtered_match_data.get("CourtSchedules", []) or [])

    return {
        "changed_court_count": court_count,
        "changed_court_names": sorted(changed_court_names, key=natural_sort_key),
        "assignment_filename": f"ChangedCourts_AssignmentGrid_{event_name_no_spaces}_{date_suffix}{suffix}.xlsx",
        "assignment_bytes": save_workbook_to_bytes(assignment_wb),
        "worksheet_filename": f"ChangedCourts_WorksheetGrid_{event_name_no_spaces}_{date_suffix}{suffix}.xlsx",
        "worksheet_bytes": save_workbook_to_bytes(worksheet_wb),
    }


# =========================
# STREAMLIT APP
# =========================

st.set_page_config(
    page_title="AES Grid Workbook Generator",
    page_icon="🏐",
    layout="wide",
)

top_left, top_right = st.columns([5, 1])

with top_left:
    st.title("AES Grid Workbook Generator")

with top_right:
    st.button("Start Over", on_click=reset_app, use_container_width=True)

st.write(
    "Paste one or more AES event keys or full AES schedule URLs, review the combined event information, "
    "set division level abbreviations, then generate both Excel workbooks."
)

if "event_info" not in st.session_state:
    st.session_state.event_info = None

if "event_key" not in st.session_state:
    st.session_state.event_key = ""

if "event_keys" not in st.session_state:
    st.session_state.event_keys = []

if "event_records" not in st.session_state:
    st.session_state.event_records = []

if "event_input_value" not in st.session_state:
    st.session_state.event_input_value = ""

if "generated_workbooks" not in st.session_state:
    st.session_state.generated_workbooks = None

if "level_rows" not in st.session_state:
    st.session_state.level_rows = None

if "event_match_counts" not in st.session_state:
    st.session_state.event_match_counts = None

if "comparison_result" not in st.session_state:
    st.session_state.comparison_result = None

if "comparison_uploaded_summary" not in st.session_state:
    st.session_state.comparison_uploaded_summary = None

if "comparison_selected_date" not in st.session_state:
    st.session_state.comparison_selected_date = None

if "comparison_changed_workbooks" not in st.session_state:
    st.session_state.comparison_changed_workbooks = None


def clear_loaded_event_state():
    clear_assignment_download_state()
    st.session_state.event_info = None
    st.session_state.event_key = ""
    st.session_state.event_keys = []
    st.session_state.event_records = []
    st.session_state.generated_workbooks = None
    st.session_state.level_rows = None
    st.session_state.event_match_counts = None
    st.session_state.comparison_result = None
    st.session_state.comparison_uploaded_summary = None
    st.session_state.comparison_selected_date = None
    st.session_state.comparison_changed_workbooks = None


def load_events_from_input(raw_input: str, source_label: str = "event set"):
    clean_keys = extract_event_keys(raw_input)

    if not clean_keys:
        st.error("Enter at least one event key or URL.")
        return

    try:
        clear_assignment_download_state()
        with st.spinner(f"Loading {source_label}..."):
            event_records = build_event_records(clean_keys)
            combined_event_info = build_combined_event_info(event_records)

            st.session_state.event_records = event_records
            st.session_state.event_keys = clean_keys
            st.session_state.event_key = clean_keys[0] if len(clean_keys) == 1 else "|".join(clean_keys)
            st.session_state.event_info = combined_event_info
            st.session_state.event_input_value = raw_input
            st.session_state.generated_workbooks = None
            st.session_state.comparison_result = None
            st.session_state.comparison_uploaded_summary = None
            st.session_state.comparison_selected_date = None
            st.session_state.comparison_changed_workbooks = None
            st.session_state.level_rows = build_level_rows(
                combined_event_info.get("Divisions", [])
            )
            st.session_state.event_match_counts = build_event_match_count_summary(
                event_records,
                combined_event_info,
            )

        if len(clean_keys) == 1:
            st.success(f"Loaded {source_label}.")
        else:
            st.success(f"Loaded {len(clean_keys)} events into one combined grid.")
    except Exception as exc:
        clear_loaded_event_state()
        st.error(f"Could not load {source_label}: {exc}")


def load_event_from_input(raw_input: str, source_label: str = "event"):
    """Backward-compatible wrapper for older calls."""
    load_events_from_input(raw_input, source_label)


def selected_assignment_role_ids_from_checkboxes(prefix: str = "assignment_row_include") -> list[str]:
    selected_role_ids = []

    st.write(
        "Choose which rows should be included in the Assignment Grid. "
        "All rows are checked by default."
    )

    checkbox_columns = st.columns(2)

    for index, option in enumerate(ASSIGNMENT_ROLE_OPTIONS):
        key = f"{prefix}_{option['id']}"
        with checkbox_columns[index % 2]:
            is_selected = st.checkbox(
                option["checkbox_label"],
                value=True,
                key=key,
            )

        if is_selected:
            selected_role_ids.append(option["id"])

    return selected_role_ids


def assignment_download_signature(
    event_records: list[dict],
    division_settings: dict,
    selected_role_ids: list[str],
) -> tuple:
    division_signature = tuple(
        sorted(
            (
                str(division_id),
                str(settings.get("abbreviation", "")),
                str(settings.get("color", "")),
            )
            for division_id, settings in (division_settings or {}).items()
        )
    )

    return event_record_key_signature(event_records), tuple(selected_role_ids), division_signature


def render_assignment_download_options(
    event_records: list[dict],
    event_info: dict,
    division_settings: dict,
):
    selected_role_ids = selected_assignment_role_ids_from_checkboxes()
    signature = assignment_download_signature(
        event_records,
        division_settings,
        selected_role_ids,
    )

    if not selected_role_ids:
        st.error("Select at least one row before preparing the download.")
        prepare_disabled = True
    else:
        selected_labels = [
            option["checkbox_label"]
            for option in ASSIGNMENT_ROLE_OPTIONS
            if option["id"] in set(selected_role_ids)
        ]
        st.caption(f"Rows selected: {', '.join(selected_labels)}")
        prepare_disabled = False

    prepare_clicked = st.button(
        "Prepare Assignment Grid Download",
        type="primary",
        disabled=prepare_disabled,
        key="prepare_custom_assignment_download",
        use_container_width=True,
    )

    if prepare_clicked:
        try:
            with st.spinner("Preparing Assignment Grid download..."):
                custom_file = create_assignment_grid_file(
                    event_records=event_records,
                    event_info=event_info,
                    division_settings=division_settings,
                    assignment_role_ids=selected_role_ids,
                )

            st.session_state.assignment_download_custom = {
                "signature": signature,
                "role_ids": selected_role_ids,
                **custom_file,
            }
            st.success("Assignment Grid is ready to download.")
        except Exception as exc:
            st.session_state.assignment_download_custom = None
            st.error(f"Could not prepare Assignment Grid download: {exc}")

    prepared_download = st.session_state.get("assignment_download_custom")

    if prepared_download and prepared_download.get("signature") == signature:
        st.download_button(
            label="Download Assignment Grid",
            data=prepared_download["assignment_bytes"],
            file_name=prepared_download["assignment_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="download_custom_assignment_grid",
            use_container_width=True,
        )
    elif prepared_download:
        st.info("Selections changed. Prepare the download again before downloading.")

    if st.button("Close", key="close_assignment_download_options", use_container_width=True):
        st.session_state.show_assignment_download_options = False


def preset_event_urls(preset_event: dict) -> list[str]:
    if "urls" in preset_event:
        return [str(url).strip() for url in preset_event.get("urls", []) if str(url).strip()]

    url = str(preset_event.get("url", "")).strip()
    return [url] if url else []


def preset_event_is_active(preset_event: dict) -> bool:
    return bool(preset_event_urls(preset_event)) and bool(preset_event.get("active", True))


def render_preset_event_buttons():
    st.markdown("#### Quick Event Links")

    for row_index, preset_row in enumerate(PRESET_EVENT_BUTTON_ROWS):
        columns = st.columns(len(preset_row))

        for column, preset_event in zip(columns, preset_row):
            label = preset_event.get("label", "Event")
            preset_input = event_key_input_text(preset_event_urls(preset_event))
            is_active = preset_event_is_active(preset_event)
            help_text = "Load this preset event set." if is_active else "No active link is configured for this button."

            with column:
                clicked = st.button(
                    label,
                    key=f"preset_event_{row_index}_{label}",
                    use_container_width=True,
                    disabled=not is_active,
                    help=help_text,
                )

            if clicked:
                st.session_state.event_input_value = preset_input
                load_events_from_input(preset_input, label)


render_preset_event_buttons()
st.divider()

event_input = st.text_area(
    "AES Event Key(s) or Schedule URL(s)",
    placeholder=(
        "Paste one event per line, or separate multiple events with commas.\n"
        "Example: PTAwMDAwNDEzMjQ90\n"
        "https://results.advancedeventsystems.com/event/PTAwMDAwNDEzMjQ90/home"
    ),
    key="event_input_value",
    height=120,
)

fetch_clicked = st.button("Fetch Event(s)", type="primary")

if fetch_clicked:
    load_events_from_input(event_input, "event set")


event_info = st.session_state.event_info

if event_info:
    event_records = st.session_state.event_records
    event_keys = st.session_state.event_keys
    event_name = event_info.get("DisplayName") or event_info.get("Name", "")
    location = event_info.get("Location", "")
    start_date = parse_event_date(event_info["StartDate"])
    end_date = parse_event_date(event_info["EndDate"])

    st.subheader("Event Information")
    st.markdown(f"### {event_name}")

    c1, c2, c3 = st.columns([1.3, 1.3, 2.8])

    c1.metric("Start Date", format_long_date(start_date))
    c2.metric("End Date", format_long_date(end_date))
    c3.metric("Location", location or "Not listed")

    if len(event_keys) <= 1:
        st.caption(f"AES event key: `{event_keys[0] if event_keys else st.session_state.event_key}`")
    else:
        st.caption(f"AES event keys: `{', '.join(event_keys)}`")
        with st.expander("Events included in this combined grid"):
            for record in event_records:
                record_info = record.get("event_info", {}) or {}
                st.write(f"- **{record_info.get('Name', record.get('event_key'))}** — `{record.get('event_key')}`")

    if st.session_state.event_match_counts is None:
        try:
            with st.spinner("Counting matches by day..."):
                st.session_state.event_match_counts = build_event_match_count_summary(
                    event_records,
                    event_info,
                )
        except Exception as exc:
            st.warning(f"Could not load match counts: {exc}")

    render_event_match_counts(st.session_state.event_match_counts)

    st.subheader("Division Level Abbreviations")

    st.write(
        "Set one abbreviation for each unique division level. "
        "The app will add the age automatically. For example, Classic → CL gives 11CL, 12CL, 13CL. Girls → U."
    )

    if st.session_state.level_rows is None:
        st.session_state.level_rows = build_level_rows(event_info.get("Divisions", []))

    edited_level_rows = st.data_editor(
        st.session_state.level_rows,
        use_container_width=True,
        num_rows="fixed",
        hide_index=True,
        column_order=[
            "Division Level",
            "Abbreviation",
            "Divisions Using This Level",
        ],
        column_config={
            "Division Level": st.column_config.TextColumn(
                "Division Level",
                disabled=True,
                width="medium",
            ),
            "Abbreviation": st.column_config.TextColumn(
                "Abbreviation",
                help="Example: Open → O, Premier → P, Classic → CL, Aspire → AS, Girls → U",
                width="small",
            ),
            "Divisions Using This Level": st.column_config.TextColumn(
                "Divisions Using This Level",
                disabled=True,
                width="large",
            ),
        },
        key="level_editor",
    )

    level_abbreviations = resolve_level_abbreviations(edited_level_rows)

    division_settings = build_division_settings_from_levels(
        divisions=event_info.get("Divisions", []),
        level_abbreviations=level_abbreviations,
    )

    st.subheader("Generate Workbooks")

    generate_clicked = st.button("Generate", type="primary")

    if generate_clicked:
        try:
            clear_assignment_download_state()
            with st.spinner("Generating..."):
                st.session_state.generated_workbooks = create_workbooks(
                    event_records=event_records,
                    event_info=event_info,
                    division_settings=division_settings,
                )

            st.success("Done.")
        except Exception as exc:
            st.session_state.generated_workbooks = None
            st.error(f"Could not generate workbooks: {exc}")

    generated = st.session_state.generated_workbooks

    if generated:
        st.markdown("#### Downloads")

        d1, d2 = st.columns(2)

        with d1:
            assignment_options_clicked = st.button(
                "Download Assignment Grid",
                type="primary",
                key="open_assignment_download_options",
                use_container_width=True,
            )

            if assignment_options_clicked:
                st.session_state.show_assignment_download_options = True
                st.session_state.assignment_download_custom = None

        with d2:
            st.download_button(
                label="Download Worksheet Grid",
                data=generated["worksheet_bytes"],
                file_name=generated["worksheet_filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="download_worksheet_grid",
                use_container_width=True,
            )

        if st.session_state.get("show_assignment_download_options"):
            dialog_decorator = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)

            if dialog_decorator:
                @dialog_decorator("Assignment Grid Download Options")
                def assignment_download_dialog():
                    render_assignment_download_options(
                        event_records=event_records,
                        event_info=event_info,
                        division_settings=division_settings,
                    )

                assignment_download_dialog()
            else:
                with st.expander("Assignment Grid Download Options", expanded=True):
                    render_assignment_download_options(
                        event_records=event_records,
                        event_info=event_info,
                        division_settings=division_settings,
                    )


    st.divider()
    st.subheader("Compare Uploaded Workbook to Online Schedule")
    st.write(
        "Upload a previously generated or manually edited grid workbook, choose the day/sheet, "
        "and compare its court/time match structure to the current combined AES online schedule."
    )

    available_compare_dates = list(date_range(start_date, end_date))

    selected_compare_date = st.selectbox(
        "Day to compare",
        options=available_compare_dates,
        format_func=format_long_date,
        key="compare_date",
    )

    uploaded_compare_workbook = st.file_uploader(
        "Upload workbook to compare",
        type=["xlsx", "xlsm"],
        key="compare_workbook",
    )

    selected_compare_sheet = None

    if uploaded_compare_workbook is not None:
        try:
            compare_sheet_names = get_uploaded_workbook_sheet_names(uploaded_compare_workbook)
            default_sheet_index = default_uploaded_sheet_index(
                compare_sheet_names,
                selected_compare_date,
            )

            selected_compare_sheet = st.selectbox(
                "Workbook sheet to read",
                options=compare_sheet_names,
                index=default_sheet_index,
                key="compare_sheet",
            )
        except Exception as exc:
            st.error(f"Could not read workbook sheets: {exc}")

    compare_clicked = st.button(
        "Compare Uploaded Workbook to Current Online Schedule",
        type="secondary",
        key="compare_schedule_button",
    )

    if compare_clicked:
        if uploaded_compare_workbook is None:
            st.error("Upload a workbook first.")
        elif not selected_compare_sheet:
            st.error("Select a worksheet from the uploaded workbook.")
        else:
            try:
                with st.spinner("Comparing workbook to current AES schedule..."):
                    uploaded_summary = parse_uploaded_workbook_schedule(
                        uploaded_compare_workbook,
                        selected_compare_sheet,
                    )
                    current_match_data = get_combined_match_info(
                        event_records,
                        selected_compare_date,
                    )
                    online_entries = build_online_schedule_entries(
                        current_match_data,
                        TIMEZONE,
                        division_settings,
                    )
                    comparison_result = compare_schedule_structures(
                        uploaded_summary["entries"],
                        online_entries,
                    )

                    st.session_state.comparison_result = comparison_result
                    st.session_state.comparison_uploaded_summary = uploaded_summary
                    st.session_state.comparison_selected_date = selected_compare_date
                    st.session_state.comparison_changed_workbooks = None

                    changed_court_names = changed_court_names_from_comparison(comparison_result)

                    if changed_court_names:
                        st.session_state.comparison_changed_workbooks = create_changed_court_workbooks(
                            event_info=event_info,
                            sheet_date=selected_compare_date,
                            match_data=current_match_data,
                            division_settings=division_settings,
                            changed_court_names=changed_court_names,
                        )

            except Exception as exc:
                st.session_state.comparison_result = None
                st.session_state.comparison_uploaded_summary = None
                st.session_state.comparison_selected_date = None
                st.session_state.comparison_changed_workbooks = None
                st.error(f"Could not compare schedules: {exc}")

    if (
        st.session_state.comparison_result is not None
        and st.session_state.comparison_uploaded_summary is not None
        and st.session_state.comparison_selected_date is not None
    ):
        st.markdown("#### Comparison Results")
        render_comparison_results(
            st.session_state.comparison_result,
            st.session_state.comparison_uploaded_summary,
            st.session_state.comparison_selected_date,
        )

    changed_downloads = st.session_state.comparison_changed_workbooks

    if changed_downloads:
        st.markdown("#### Changed-Court Downloads")
        st.caption(
            f"These files include only the {changed_downloads['changed_court_count']} court(s) with detected changes."
        )

        cd1, cd2 = st.columns(2)

        with cd1:
            st.download_button(
                label="Download Changed Courts Assignment Grid",
                data=changed_downloads["assignment_bytes"],
                file_name=changed_downloads["assignment_filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="download_changed_assignment_grid",
            )

        with cd2:
            st.download_button(
                label="Download Changed Courts Worksheet Grid",
                data=changed_downloads["worksheet_bytes"],
                file_name=changed_downloads["worksheet_filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="download_changed_worksheet_grid",
            )

else:
    st.info("Enter one or more AES event keys or full AES schedule URLs to begin.")